import sqlite3
import xlrd
import openpyxl
import os

'''
@Author : h31i
@Date   : 2019/9/7
@Python version: 3.7 

# ##----------------------------------------------##

1. sqlite3是支持中文做数据库名称和字段名的
2. openpyxl操作07版以上的xlsx文件
3. xlrd,xlwt分别用于读取和写xls文件

# ##----------------------------------------------##

目标：

1. excel文件名作为数据库名称，或者自定义输入名称
2. 一个sheet一个表表名为sheet1，sheet2 ... sheet255
3. 判别文件类型，导入对应操作包，如：如果是xls文件，则导入xlrd。如果是xlsx文件，导入openpyxl
4. 该程序仅限于一对一，不能进行一对多生成（一对多生成以后可以考虑尝试）

'''


class ExcelToSqlite(object):
    
    def __init__(self, excel_name):
        
        # 获取文件信息
        self.file_path = os.path.abspath(excel_name)
        self.file_name = os.path.basename(excel_name)
        self.pre_name, self.suf_name = os.path.splitext(file_name)

        # 打印文件信息
        self._file_info()
        # print("------->"+self.pre_name+'.db')
        
        print('#'*30)
        print("###--初始化数据库")
        print('#'*30)

        self.conn = sqlite3.connect(self.pre_name+'.db')
        
        if self.suf_name == ".xls":
            self.excel_object = xlrd.open_workbook(excel_name)
        elif self.suf_name == ".xlsx":
            self.excel_object = openpyxl.load_workbook(excel_name)
        else:
            print("请查看文件是否符合要求")
        
        self.cursor = self.conn.cursor()

        # super(ExcelToDB, self).__init__()

    def __del__(self):
        
        print("###--释放数据库")
        self.cursor.close()
        self.conn.close()

    def _file_info(self):

    	print("###--文件信息")
    	print("###--文件--路径："+ self.file_path)
    	print("###--文件--名称："+ self.file_name)
    	print("###--文件前缀名："+ self.pre_name)
    	print("###--文件后缀名："+ self.suf_name)

    def _sheet_info(self):
        # 格式输出
        print("###--sheet info")
        print('###--工作表名称为:', self.sheet_name)
        print('###--工作表行数为:', self.sheet_row)
        print('###--工作表列数为:', self.sheet_col)
        print('###--工作表  表头:', self.table_headers)

    # xls 版
    def xls_sheet_info(self, sheet_index):

        # 获取指定sheet对象
        self.sheet_name = self.excel_object.sheet_names()[sheet_index]
        self.sheet_object = self.excel_object.sheets()[sheet_index]

        # 获取sheet行数列数
        self.sheet_row = self.sheet_object.nrows
        self.sheet_col = self.sheet_object.ncols

        # 获取字段名
        self.table_headers = self.sheet_object.row_values(0)

        self._sheet_info()

    # xlsx 版
    def xlsx_sheet_info(self, sheet_index):

        # 获取指定sheet对象
        self.sheet_name = self.excel_object.sheetnames[sheet_index]
        self.sheet_object = self.excel_object[self.sheet_name]

        # 获取sheet行数列数
        self.sheet_row = self.sheet_object.max_row
        self.sheet_col = self.sheet_object.max_column
        
        # openpyxl 读写单元格的坐标位置起始值是(1,1)
        # 获取字段名
        self.table_headers = [self.sheet_object.cell(row=1,column=i).value for i in range(1,self.sheet_col+1)]
        
        if None in self.table_headers:
        	self.table_headers = [i for i in self.table_headers if i != None]
        	self.sheet_col = len(self.table_headers)
        
        self._sheet_info()


    # 输出指定区间的数据xls版
    def show_data(self,head=1,end=10):
        
        for row in range(head,end):
            print(self.sheet_object.row_values(row))


    # 生成建表SQL,中文表名中文字段
    def table_sql(self):
        '''
        create table if not exists table_name(
            name text,
            date text
        );
        暂时将格式都做成text类型
        '''
        sql = "create table if not exists '%s'(\n\'%s);" % (self.sheet_name,'\' text,\n\''.join(self.table_headers) + "' text")
        
        self.cursor.execute(sql)

        print("###--根据sheet名创建表")
        print('#'*30)
        print(sql)
        print('#'*30)


    # xls类型插入数据
    def xls_insert_data(self, log=False):

        print("###--日志记录："+str(log))
        # for row_index in range(1,10):
        for row_index in range(1,self.sheet_row):
            sql = "insert into '%s'%svalues%s;" % (self.sheet_name, str(tuple(self.table_headers)),str(tuple(self.sheet_object.row_values(row_index))) )
            self.cursor.execute(sql)
            
            if log:
            	print(sql)
        
        print("提交事务")
        self.conn.commit()

    
    # xlsx类型插入数据
    # openpyxl value获取值如果为空返回None（真够麻烦）
    def xlsx_insert_data(self, log=False):

        print("###--日志记录："+str(log))
        
        for row_index in range(2,self.sheet_row+1):

            row_data = [self.sheet_object.cell(row=row_index,column=col_index).value for col_index in range(1,self.sheet_col+1)]
            row_data = [l if l is not None else '' for l in row_data]
            
            # print(row_data)
            sql = "insert into '%s'%svalues%s;" % (self.sheet_name, str(tuple(self.table_headers)),str(tuple(row_data)) )
            
            if log:
            	print(sql)

            self.cursor.execute(sql)
        
        print("###--提交事务")
        self.conn.commit()


if __name__ == '__main__':

	# 文件名称
    file_name = ".\\beijiao.xls"

    # 创建对象
    excelToSqlite = ExcelToSqlite(file_name)

    # excelToSqlite.xls_sheet_info(0)
    excelToSqlite.xlsx_sheet_info(0)

    # 目前仅对xls文件有效
    # excelToSqlite.show_data()
    # excelToSqlite.show_data(1900,1904)

    excelToSqlite.table_sql()

    # excelToSqlite.xls_insert_data()
    excelToSqlite.xlsx_insert_data(log=True)